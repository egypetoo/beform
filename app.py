from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path
from threading import Lock
import csv
import hmac
import io
import json
import os
import secrets
import time
import uuid

from dotenv import load_dotenv
from flask import Flask, Response, flash, redirect, render_template, request, send_from_directory, session, url_for
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash
import requests

import user_store

SHEET_SESSION = requests.Session()

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")
GOOGLE_SHEET_WEBHOOK = os.getenv("GOOGLE_SHEET_WEBHOOK", "").strip()
MANAGER_LOGIN_PATH = os.getenv("MANAGER_LOGIN_PATH", "/be-review-k4n").strip() or "/be-review-k4n"
if not MANAGER_LOGIN_PATH.startswith("/"):
    MANAGER_LOGIN_PATH = "/" + MANAGER_LOGIN_PATH

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1)
app.secret_key = os.getenv("FLASK_SECRET_KEY") or secrets.token_hex(32)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.getenv("SESSION_COOKIE_SECURE", "1") != "0",
    PERMANENT_SESSION_LIFETIME=timedelta(hours=8),
)

ROWS_CACHE = {}
CACHE_SECONDS = 60
TRACK_CACHE = {}
TRACK_CACHE_SECONDS = 60
RECENT_SUBMISSIONS = {}
RECENT_LOCK = Lock()
DEDUP_SECONDS = 90
LOGIN_ATTEMPTS = {}
LOGIN_LOCK = Lock()
MAX_LOGIN_ATTEMPTS = 5
LOGIN_WINDOW_SECONDS = 300
TRACK_ATTEMPTS = {}
TRACK_LOCK = Lock()
MAX_TRACK_ATTEMPTS = 8
TRACK_WINDOW_SECONDS = 300


def today_values():
    now = datetime.now()
    return {
        "today": now.strftime("%Y-%m-%d"),
        "today_display": now.strftime("%A, %B %d, %Y"),
    }


def shift_month(year: int, month: int, delta: int) -> tuple[int, int]:
    month += delta
    year += (month - 1) // 12
    month = (month - 1) % 12 + 1
    return year, month


def cycle_start_for(day) -> datetime:
    if day.day >= 25:
        return datetime(day.year, day.month, 25)
    year, month = shift_month(day.year, day.month, -1)
    return datetime(year, month, 25)


def cycle_end_for(start: datetime) -> datetime:
    year, month = shift_month(start.year, start.month, 1)
    return datetime(year, month, 24)


def payroll_cycles(count: int = 8) -> list:
    start = cycle_start_for(datetime.now())
    cycles = []
    for _ in range(count):
        end = cycle_end_for(start)
        cycles.append({
            "value": start.strftime("%Y-%m-%d"),
            "label": f"{start.strftime('%d %b')} – {end.strftime('%d %b %Y')}",
            "start": start.strftime("%Y-%m-%d"),
            "end": end.strftime("%Y-%m-%d"),
        })
        year, month = shift_month(start.year, start.month, -1)
        start = datetime(year, month, 25)
    return cycles


def row_matches_date_range(row: dict, date_from: str, date_to: str) -> bool:
    submitted = str(row.get("Submitted At") or "")[:10]
    start = str(row.get("From Date") or "")[:10] or submitted
    end = str(row.get("To Date") or "")[:10] or start
    if date_from and end and end < date_from:
        return False
    if date_to and start and start > date_to:
        return False
    return True


def current_cycle_value() -> str:
    return cycle_start_for(datetime.now()).strftime("%Y-%m-%d")

LEAVE_GROUPS = [
    {
        "title": "Work",
        "title_ar": "العمل",
        "options": [
            {"value": "work_remotely", "en": "Work Remotely", "ar": "عمل عن بعد"},
            {"value": "monthly_saturday", "en": "Monthly Saturday Work", "ar": "عمل السبت الشهري"},
        ],
    },
    {
        "title": "Leaves",
        "title_ar": "الإجازات",
        "options": [
            {"value": "business_mission", "en": "Business Mission", "ar": "مهمة عمل"},
            {"value": "sick_leave", "en": "Sick Leave", "ar": "إجازة مرضية"},
            {"value": "personal_excuse", "en": "Personal Excuse", "ar": "عذر شخصي"},
            {"value": "unpaid_leave", "en": "Unpaid Leave", "ar": "إجازة بدون راتب"},
            {"value": "missing_punch_in", "en": "Missing Punch In", "ar": "نسيان بصمة حضور"},
            {"value": "missing_punch_out", "en": "Missing Punch Out", "ar": "نسيان بصمة انصراف"},
        ],
    },
    {
        "title": "Vacations",
        "title_ar": "الإجازات السنوية / المرضية",
        "options": [
            {"value": "annual_vacation", "en": "Annual Vacation", "ar": "إجازة سنوية"},
            {"value": "sickness_vacation", "en": "Sickness Vacation", "ar": "إجازة مرضية طويلة"},
        ],
    },
]

PUNCH_TYPES = {"missing_punch_in", "missing_punch_out"}
TIME_RANGE_TYPES = {"business_mission", "personal_excuse"}
DATE_RANGE_TYPES = {"sick_leave", "unpaid_leave", "annual_vacation", "sickness_vacation", "work_remotely"}
SATURDAY_TYPES = {"monthly_saturday"}

def all_departments(active_only: bool = True) -> list:
    return [
        {"value": dept["value"], "label": dept["label"]}
        for dept in user_store.list_custom_departments(active_only=active_only)
    ]


def department_maps():
    departments = all_departments(active_only=True)
    return {
        "items": departments,
        "values": {item["value"] for item in departments},
        "labels": {item["value"]: item["label"] for item in departments},
        "label_set": {item["label"] for item in departments},
    }


def get_managers():
    load_dotenv(BASE_DIR / ".env", override=True)
    return {
        "hr": {
            "name": "HR Manager",
            "department": "ALL",
            "role": "hr",
            "team": "",
            "password": os.getenv("MANAGER_HR_PASSWORD", ""),
        },
    }


def option_lookup():
    mapping = {}
    for group in LEAVE_GROUPS:
        for option in group["options"]:
            mapping[option["value"]] = option
    return mapping


def webhook_url():
    return os.getenv("GOOGLE_SHEET_WEBHOOK", "").strip() or GOOGLE_SHEET_WEBHOOK


def sheet_secret():
    return os.getenv("SHEET_SECRET", "").strip()


def password_matches(stored: str, provided: str) -> bool:
    if not stored or not provided:
        return False
    if stored.startswith(("pbkdf2:", "scrypt:", "argon2:")):
        return check_password_hash(stored, provided)
    return hmac.compare_digest(stored, provided)


def get_csrf_token() -> str:
    token = session.get("csrf_token")
    if not token:
        token = secrets.token_hex(32)
        session["csrf_token"] = token
    return token


def csrf_is_valid() -> bool:
    sent = request.form.get("csrf_token", "")
    expected = session.get("csrf_token", "")
    return bool(sent) and bool(expected) and hmac.compare_digest(sent, expected)


def client_ip() -> str:
    return request.remote_addr or "unknown"


def login_is_locked(ip: str) -> bool:
    now = time.time()
    with LOGIN_LOCK:
        record = LOGIN_ATTEMPTS.get(ip)
        if not record:
            return False
        if now - record["start"] > LOGIN_WINDOW_SECONDS:
            LOGIN_ATTEMPTS.pop(ip, None)
            return False
        return record["count"] >= MAX_LOGIN_ATTEMPTS


def record_login_failure(ip: str) -> None:
    now = time.time()
    with LOGIN_LOCK:
        record = LOGIN_ATTEMPTS.get(ip)
        if not record or now - record["start"] > LOGIN_WINDOW_SECONDS:
            LOGIN_ATTEMPTS[ip] = {"count": 1, "start": now}
        else:
            record["count"] += 1


def clear_login_failures(ip: str) -> None:
    with LOGIN_LOCK:
        LOGIN_ATTEMPTS.pop(ip, None)


def dashboard_redirect_args(form=None) -> dict:
    source = form if form is not None else request.form
    return {
        "status": source.get("status_filter") or source.get("status") or "Pending",
        "q": source.get("q", ""),
        "type": source.get("type_filter") or source.get("type") or "",
        "from": source.get("date_from") or source.get("from") or "",
        "to": source.get("date_to") or source.get("to") or "",
    }


def can_review_department(manager: dict, department: str) -> bool:
    if is_hr(manager):
        return True
    return department == manager.get("department")


def is_hr(manager: dict) -> bool:
    return manager.get("role") == "hr" or manager.get("department") == "ALL"


def manager_role(manager: dict) -> str:
    if manager.get("role"):
        return manager["role"]
    return "hr" if manager.get("department") == "ALL" else "department"


def can_review_row(manager: dict, row: dict) -> bool:
    department = str(row.get("Department") or "").strip()
    if not can_review_department(manager, department):
        return False
    if manager_role(manager) != "team":
        return True
    return str(row.get("Team") or "").strip().lower() == str(manager.get("team") or "").strip().lower()


def filter_visible_rows(manager: dict, rows: list) -> list:
    return [row for row in rows if can_review_row(manager, row)]


def teams_for_form() -> dict:
    by_label = user_store.teams_by_department()
    return {
        item["value"]: by_label.get(item["label"], [])
        for item in all_departments()
    }


def index_context(form) -> dict:
    return {
        "leave_groups": LEAVE_GROUPS,
        "departments": all_departments(),
        "teams_by_department": teams_for_form(),
        "form": form,
        **today_values(),
    }


def track_is_limited(ip: str) -> bool:
    now = time.time()
    with TRACK_LOCK:
        record = TRACK_ATTEMPTS.get(ip)
        if not record or now - record["start"] > TRACK_WINDOW_SECONDS:
            TRACK_ATTEMPTS[ip] = {"count": 1, "start": now}
            return False
        record["count"] += 1
        return record["count"] > MAX_TRACK_ATTEMPTS


@app.context_processor
def inject_security():
    manager = session.get("manager")
    return {
        "csrf_token": get_csrf_token(),
        "is_hr": bool(manager and is_hr(manager)),
    }


def sheet_api(payload: dict) -> dict:
    webhook = webhook_url()
    if not webhook:
        raise RuntimeError("Google Sheet is not connected")

    payload = dict(payload)
    payload["secret"] = sheet_secret()
    if not payload["secret"]:
        raise RuntimeError("Sheet secret is not configured")

    response = SHEET_SESSION.post(webhook, json=payload, timeout=20)
    response.raise_for_status()
    try:
        data = json.loads(response.text or "{}")
    except json.JSONDecodeError as exc:
        raise RuntimeError("Invalid sheet response") from exc
    if not data.get("ok"):
        raise RuntimeError(data.get("error") or "Google Sheet did not confirm the save")
    return data


@app.after_request
def set_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    if request.path.startswith("/static/"):
        response.headers["Cache-Control"] = "public, max-age=604800"
    return response


def save_submission(row: dict) -> dict:
    payload = dict(row)
    payload["action"] = "create"
    data = sheet_api(payload)
    ROWS_CACHE.clear()
    return data


def submission_fingerprint(row: dict) -> str:
    return "|".join([
        str(row.get("fingerprint_id") or ""),
        str(row.get("department") or ""),
        str(row.get("request_type") or ""),
        str(row.get("start_date") or ""),
        str(row.get("end_date") or ""),
        str(row.get("from_time") or ""),
        str(row.get("to_time") or ""),
        str(row.get("punch_in_time") or ""),
        str(row.get("punch_out_time") or ""),
    ])


def claim_submission(key: str) -> bool:
    now = time.time()
    with RECENT_LOCK:
        expired = [item for item, stamp in RECENT_SUBMISSIONS.items() if now - stamp > DEDUP_SECONDS]
        for item in expired:
            RECENT_SUBMISSIONS.pop(item, None)
        if key in RECENT_SUBMISSIONS:
            return False
        RECENT_SUBMISSIONS[key] = now
        return True


def list_requests(department: str) -> list:
    now = time.time()
    cached = ROWS_CACHE.get(department)
    if cached and now - cached["at"] < CACHE_SECONDS:
        return cached["rows"]

    data = sheet_api({"action": "list", "department": department})
    rows = data.get("rows", [])
    ROWS_CACHE[department] = {"at": now, "rows": rows}
    return rows


def set_request_status(items: list, status: str, reviewed_by: str, reason: str = "") -> None:
    sheet_api(
        {
            "action": "set_status",
            "items": items,
            "status": status,
            "reviewed_by": reviewed_by,
            "reason": reason,
        }
    )
    ROWS_CACHE.clear()


def normalize_fingerprint(value) -> str:
    text = str(value or "").strip()
    if text.endswith(".0") and text[:-2].replace(".", "").isdigit():
        text = text[:-2]
    return text


TRACK_LIMIT = 30


def row_submitted_key(row: dict) -> str:
    return str(row.get("Submitted At") or row.get("From Date") or "")


def lookup_by_fingerprint(fingerprint: str) -> list:
    now = time.time()
    cached = TRACK_CACHE.get(fingerprint)
    if cached and now - cached["at"] < TRACK_CACHE_SECONDS:
        return list(cached["rows"])

    try:
        data = sheet_api({
            "action": "lookup",
            "fingerprint_id": fingerprint,
            "limit": TRACK_LIMIT,
        })
        source_rows = data.get("rows", [])
    except Exception:
        data = sheet_api({"action": "list", "department": "ALL"})
        source_rows = data.get("rows", [])

    wanted = normalize_fingerprint(fingerprint)
    rows = [
        row for row in source_rows
        if normalize_fingerprint(row.get("Fingerprint Number")) == wanted
    ]
    rows.sort(key=row_submitted_key, reverse=True)
    rows = rows[:TRACK_LIMIT]
    TRACK_CACHE[fingerprint] = {"at": now, "rows": rows}
    return list(rows)


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("manager"):
            return redirect(url_for("login"))
        return view(*args, **kwargs)

    return wrapped


def hr_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        manager = session.get("manager")
        if not manager:
            return redirect(url_for("login"))
        if not is_hr(manager):
            flash("Only HR can manage team leaders.", "error")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)

    return wrapped


@app.route("/manifest.webmanifest")
def pwa_manifest():
    response = send_from_directory(BASE_DIR / "static", "manifest.webmanifest")
    response.headers["Content-Type"] = "application/manifest+json"
    return response


@app.route("/sw.js")
def pwa_service_worker():
    response = send_from_directory(BASE_DIR / "static", "sw.js")
    response.headers["Content-Type"] = "application/javascript"
    response.headers["Service-Worker-Allowed"] = "/"
    response.headers["Cache-Control"] = "no-cache"
    return response


@app.route("/", methods=["GET", "POST"])
def index():
    options = option_lookup()

    if request.method == "POST":
        if not csrf_is_valid():
            flash("The form expired. Please refresh and try again.", "error")
            return render_template("index.html", **index_context(request.form))

        fingerprint_id = request.form.get("fingerprint_id", "").strip()
        name = request.form.get("name", "").strip()
        department = request.form.get("department", "").strip()
        team = request.form.get("team", "").strip()
        request_type = request.form.get("request_type", "").strip()
        request_date = request.form.get("request_date", "").strip() or datetime.now().strftime("%Y-%m-%d")
        punch_in_time = request.form.get("punch_in_time", "").strip()
        punch_out_time = request.form.get("punch_out_time", "").strip()
        from_time = request.form.get("from_time", "").strip()
        to_time = request.form.get("to_time", "").strip()
        start_date = request.form.get("start_date", "").strip()
        end_date = request.form.get("end_date", "").strip()
        notes = request.form.get("notes", "").strip()

        errors = []
        if not fingerprint_id:
            errors.append("Fingerprint number is required")
        elif not fingerprint_id.isdigit():
            errors.append("Fingerprint number must contain digits only")
        if not name:
            errors.append("Name is required")
        if not department:
            errors.append("Department is required")
        elif department not in department_maps()["values"]:
            errors.append("Please select a valid department")
        else:
            allowed_teams = [item["value"] for item in teams_for_form().get(department, [])]
            if allowed_teams and team not in allowed_teams:
                errors.append("Please select your team leader")
            if not allowed_teams:
                team = ""
        if request_type not in options:
            errors.append("Request type is required")
        if request_type == "missing_punch_in" and not punch_in_time:
            errors.append("Actual punch-in time is required")
        if request_type == "missing_punch_out" and not punch_out_time:
            errors.append("Actual punch-out time is required")
        if request_type in SATURDAY_TYPES and start_date:
            end_date = start_date
        if request_type in TIME_RANGE_TYPES:
            if not from_time:
                errors.append("From time is required")
            if not to_time:
                errors.append("To time is required")
        if request_type in options:
            if not start_date:
                errors.append("Saturday date is required" if request_type in SATURDAY_TYPES else "From date is required")
            if request_type not in SATURDAY_TYPES and not end_date:
                errors.append("To date is required")
        if start_date and end_date and start_date > end_date:
            errors.append("From date cannot be after To date")
        if request_type in PUNCH_TYPES:
            today = datetime.now().strftime("%Y-%m-%d")
            if (start_date and start_date > today) or (end_date and end_date > today):
                errors.append("Missing punch cannot be submitted for a future date")
        if request_type in SATURDAY_TYPES and start_date:
            try:
                saturday = datetime.strptime(start_date, "%Y-%m-%d")
            except ValueError:
                saturday = None
                errors.append("Please choose a valid Saturday date")
            if saturday:
                if saturday.weekday() == 4:
                    errors.append("Friday is an off day. Choose the working Saturday instead.")
                elif saturday.weekday() != 5:
                    errors.append("Monthly Saturday Work must be a Saturday. Friday and Saturday are off except one Saturday per month.")

        if errors:
            for error in errors:
                flash(error, "error")
            return render_template("index.html", **index_context(request.form))

        selected = options[request_type]
        department_label = department_maps()["labels"].get(department, "")
        row = {
            "request_id": uuid.uuid4().hex[:12].upper(),
            "submitted_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "fingerprint_id": fingerprint_id,
            "name": name,
            "department": department_label,
            "request_type": selected["en"],
            "request_date": request_date,
            "punch_in_time": punch_in_time if request_type == "missing_punch_in" else "",
            "punch_out_time": punch_out_time if request_type == "missing_punch_out" else "",
            "from_time": from_time if request_type in TIME_RANGE_TYPES else "",
            "to_time": to_time if request_type in TIME_RANGE_TYPES else "",
            "start_date": start_date,
            "end_date": start_date if request_type in SATURDAY_TYPES else end_date,
            "notes": notes,
            "status": "Pending",
            "team": team,
        }
        if not claim_submission(submission_fingerprint(row)):
            flash("This request was already submitted.", "error")
            return render_template("index.html", **index_context(request.form))

        try:
            result = save_submission(row)
        except Exception as exc:
            (BASE_DIR / "sheet_error.log").write_text(str(exc), encoding="utf-8")
            flash("Could not save the request to the HR sheet. Please try again.", "error")
            return render_template("index.html", **index_context(request.form))

        if result.get("saturday_month"):
            flash("Only one working Saturday is allowed per month (25th to 24th). You already submitted one.", "error")
            return render_template("index.html", **index_context(request.form))

        if result.get("duplicate"):
            flash("This request was already submitted.", "error")
            return render_template("index.html", **index_context(request.form))

        if result.get("conflict"):
            if result.get("conflict_type") == "saturday":
                flash("This Saturday overlaps another leave request.", "error")
            else:
                flash("Work Remotely and Missing Punch cannot be submitted for the same day.", "error")
            return render_template("index.html", **index_context(request.form))

        return redirect(url_for("success"))

    return render_template("index.html", **index_context({}))


@app.route("/success")
def success():
    return render_template("success.html")


@app.route("/track", methods=["GET", "POST"])
def track():
    rows = None
    fingerprint_id = ""
    status_filter = "All"
    type_filter = ""
    request_types = []

    if request.method == "POST":
        if not csrf_is_valid():
            flash("The form expired. Please refresh and try again.", "error")
            return render_template(
                "track.html",
                rows=None,
                fingerprint_id="",
                status_filter="All",
                type_filter="",
                request_types=[],
                statuses=["Pending", "Approved", "Rejected", "All"],
            )

        fingerprint_id = request.form.get("fingerprint_id", "").strip()
        status_filter = request.form.get("status", "All").strip() or "All"
        type_filter = request.form.get("type", "").strip()
        if status_filter not in {"Pending", "Approved", "Rejected", "All"}:
            status_filter = "All"

        if not fingerprint_id:
            flash("Fingerprint number is required.", "error")
        elif not fingerprint_id.isdigit():
            flash("Fingerprint number must contain digits only.", "error")
        else:
            cached = TRACK_CACHE.get(fingerprint_id)
            cache_fresh = bool(cached and time.time() - cached["at"] < TRACK_CACHE_SECONDS)
            if not cache_fresh and track_is_limited(client_ip()):
                flash("Too many searches. Please wait 5 minutes and try again.", "error")
                return render_template(
                    "track.html",
                    rows=None,
                    fingerprint_id=fingerprint_id,
                    status_filter=status_filter,
                    type_filter=type_filter,
                    request_types=[],
                    statuses=["Pending", "Approved", "Rejected", "All"],
                )
            try:
                all_rows = lookup_by_fingerprint(fingerprint_id)
                if not all_rows:
                    flash("No requests found for this fingerprint.", "error")
                    rows = None
                else:
                    request_types = sorted({
                        str(row.get("Request Type") or "").strip()
                        for row in all_rows
                        if row.get("Request Type")
                    })
                    rows = all_rows
                    if status_filter != "All":
                        rows = [row for row in rows if (row.get("Status") or "Pending") == status_filter]
                    if type_filter:
                        rows = [row for row in rows if (row.get("Request Type") or "") == type_filter]
            except Exception as exc:
                (BASE_DIR / "sheet_error.log").write_text(str(exc), encoding="utf-8")
                flash("Could not load requests. Please try again.", "error")
                rows = None

    return render_template(
        "track.html",
        rows=rows,
        fingerprint_id=fingerprint_id,
        status_filter=status_filter,
        type_filter=type_filter,
        request_types=request_types,
        statuses=["Pending", "Approved", "Rejected", "All"],
    )


@app.route(MANAGER_LOGIN_PATH, methods=["GET", "POST"])
def login():
    if session.get("manager"):
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        if not csrf_is_valid():
            flash("The form expired. Please refresh and try again.", "error")
            return render_template("login.html")

        ip = client_ip()
        if login_is_locked(ip):
            flash("Too many login attempts. Please wait 5 minutes and try again.", "error")
            return render_template("login.html")

        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        manager = get_managers().get(username)
        if manager:
            if not password_matches(manager.get("password", ""), password):
                manager = None
        else:
            stored = user_store.find_user(username)
            if stored and password_matches(stored.get("password_hash", ""), password):
                manager = stored
            else:
                manager = None
        if not manager:
            record_login_failure(ip)
            flash("Invalid username or password.", "error")
            return render_template("login.html")

        clear_login_failures(ip)
        session.clear()
        session.permanent = True
        session["csrf_token"] = secrets.token_hex(32)
        session["manager"] = {
            "username": username,
            "name": manager["name"],
            "department": manager["department"],
            "role": manager.get("role") or ("hr" if manager["department"] == "ALL" else "department"),
            "team": manager.get("team") or "",
        }
        return redirect(url_for("dashboard"))

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    manager = session["manager"]
    status_filter = request.args.get("status", "Pending")
    search_query = request.args.get("q", "").strip()
    type_filter = request.args.get("type", "").strip()
    date_from = request.args.get("from", "").strip()
    date_to = request.args.get("to", "").strip()
    if len(date_from) != 10:
        date_from = ""
    if len(date_to) != 10:
        date_to = ""
    try:
        rows = list_requests(manager["department"])
        rows = filter_visible_rows(manager, rows)
    except Exception as exc:
        (BASE_DIR / "sheet_error.log").write_text(str(exc), encoding="utf-8")
        rows = []
        flash("Could not load requests from the HR sheet.", "error")

    request_types = sorted({
        str(row.get("Request Type") or "").strip()
        for row in rows
        if row.get("Request Type")
    })

    if status_filter and status_filter != "All":
        rows = [row for row in rows if (row.get("Status") or "Pending") == status_filter]
    if type_filter:
        rows = [row for row in rows if (row.get("Request Type") or "") == type_filter]
    if date_from or date_to:
        rows = [row for row in rows if row_matches_date_range(row, date_from, date_to)]
    if search_query:
        needle = search_query.lower()
        rows = [
            row for row in rows
            if needle in " ".join([
                str(row.get("Name") or ""),
                str(row.get("Fingerprint Number") or ""),
                str(row.get("Department") or ""),
                str(row.get("Team") or ""),
                str(row.get("Request Type") or ""),
                str(row.get("Notes") or ""),
                str(row.get("Rejection Reason") or ""),
            ]).lower()
        ]

    return render_template(
        "dashboard.html",
        manager=manager,
        rows=rows,
        status_filter=status_filter,
        search_query=search_query,
        type_filter=type_filter,
        date_from=date_from,
        date_to=date_to,
        request_types=request_types,
        statuses=["Pending", "Approved", "Rejected", "All"],
    )


@app.route("/dashboard/status", methods=["POST"])
@login_required
def update_status():
    manager = session["manager"]
    if not csrf_is_valid():
        flash("The form expired. Please refresh and try again.", "error")
        return redirect(url_for("dashboard"))

    status = request.form.get("status", "").strip()
    selected = request.form.getlist("selected")
    request_id = request.form.get("request_id", "").strip()
    department = request.form.get("department", "").strip() or manager["department"]

    items = []
    if selected:
        for value in selected:
            request_id_value, _, dept_value = value.partition("||")
            if request_id_value:
                items.append({
                    "request_id": request_id_value.strip(),
                    "department": dept_value.strip() or department,
                })
    elif request_id:
        items.append({"request_id": request_id, "department": department})

    if status not in {"Approved", "Rejected"} or not items:
        flash("Select at least one request first.", "error")
        return redirect(url_for("dashboard", **dashboard_redirect_args()))

    reason = request.form.get("rejection_reason", "").strip()
    if status == "Rejected":
        if not reason:
            flash("Please enter a rejection reason.", "error")
            return redirect(url_for("dashboard", **dashboard_redirect_args()))
        reason = reason[:300]
    else:
        reason = ""

    try:
        visible = filter_visible_rows(manager, list_requests(manager["department"]))
    except Exception as exc:
        (BASE_DIR / "sheet_error.log").write_text(str(exc), encoding="utf-8")
        flash("Could not verify the selected requests.", "error")
        return redirect(url_for("dashboard", **dashboard_redirect_args()))

    visible_by_id = {
        str(row.get("Request ID") or "").strip(): row
        for row in visible
        if row.get("Request ID")
    }
    authorized = []
    for item in items:
        request_id_value = item["request_id"]
        row = visible_by_id.get(request_id_value)
        if not row or not can_review_row(manager, row):
            continue
        authorized.append({
            "request_id": request_id_value,
            "department": str(row.get("Department") or "").strip(),
        })

    if not authorized:
        flash("You can only review requests assigned to you.", "error")
        return redirect(url_for("dashboard", **dashboard_redirect_args()))

    try:
        set_request_status(authorized, status, manager["name"], reason)
        flash(f"{len(authorized)} request(s) {status.lower()} successfully.", "success")
    except Exception as exc:
        (BASE_DIR / "sheet_error.log").write_text(str(exc), encoding="utf-8")
        flash("Could not update the request status.", "error")

    return redirect(url_for("dashboard", **dashboard_redirect_args()))


@app.route("/users", methods=["GET", "POST"])
@hr_required
def users_admin():
    maps = department_maps()
    if request.method == "POST":
        if not csrf_is_valid():
            flash("The form expired. Please refresh and try again.", "error")
            return redirect(url_for("users_admin"))

        username = request.form.get("username", "")
        name = request.form.get("name", "")
        department_value = request.form.get("department", "").strip()
        team = request.form.get("team", "")
        password = request.form.get("password", "")
        department_label = maps["labels"].get(department_value, "")
        errors = user_store.validate_new_user(
            username,
            name,
            department_label,
            team,
            password,
            maps["label_set"],
        )
        if errors:
            for error in errors:
                flash(error, "error")
        else:
            try:
                user_store.create_user(username, name, department_label, team, password)
                flash("Team leader added.", "success")
            except ValueError as exc:
                flash(str(exc), "error")
        return redirect(url_for("users_admin"))

    return render_template(
        "users.html",
        manager=session["manager"],
        users=user_store.list_users(),
        departments=maps["items"],
        custom_departments=user_store.list_custom_departments(active_only=False),
    )


@app.route("/users/department", methods=["POST"])
@hr_required
def users_department():
    if not csrf_is_valid():
        flash("The form expired. Please refresh and try again.", "error")
        return redirect(url_for("users_admin"))

    label = request.form.get("label", "")
    name = request.form.get("manager_name", "")
    username = request.form.get("manager_username", "")
    password = request.form.get("manager_password", "")
    errors = user_store.validate_new_department(label, name, username, password)
    if errors:
        for error in errors:
            flash(error, "error")
    else:
        try:
            created = user_store.create_department(label, name, username, password)
            flash(f"Department {created['label']} added. Manager login: {created['username']}", "success")
        except ValueError as exc:
            flash(str(exc), "error")
    return redirect(url_for("users_admin"))


IMPORT_TEMPLATE = """role,department,team,name,username,password
department,Web,,Mohamed,web,ChangeMe123
team,Web,Ahmed team,Ahmed,ahmed,ChangeMe123
team,Web,Omar team,Omar,omar,ChangeMe123
department,Marketing,,Sara,marketing,ChangeMe123
"""
MAX_IMPORT_BYTES = 200_000
MAX_IMPORT_ROWS = 200


def parse_people_file(upload) -> list:
    filename = (upload.filename or "").lower()
    raw = upload.read(MAX_IMPORT_BYTES + 1)
    if len(raw) > MAX_IMPORT_BYTES:
        raise ValueError("File is too large. Keep it under 200 KB.")
    if filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Excel .xlsx needs openpyxl. Save as CSV UTF-8 instead.") from exc
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            headers = [str(cell or "").strip().lower() for cell in next(rows_iter)]
        except StopIteration as exc:
            raise ValueError("The file is empty.") from exc
        rows = []
        for values in rows_iter:
            rows.append({headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))})
            if len(rows) > MAX_IMPORT_ROWS:
                raise ValueError("Too many rows. Import up to 200 at a time.")
        return rows
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("Save the file as CSV UTF-8.") from exc
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("The CSV has no header row.")
    rows = list(reader)
    if len(rows) > MAX_IMPORT_ROWS:
        raise ValueError("Too many rows. Import up to 200 at a time.")
    return rows


@app.route("/users/template.csv")
@hr_required
def users_template():
    return Response(
        IMPORT_TEMPLATE,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=be-people-template.csv"},
    )


@app.route("/users/import", methods=["POST"])
@hr_required
def users_import():
    if not csrf_is_valid():
        flash("The form expired. Please refresh and try again.", "error")
        return redirect(url_for("users_admin"))
    upload = request.files.get("sheet")
    if not upload or not upload.filename:
        flash("Choose a CSV or Excel file first.", "error")
        return redirect(url_for("users_admin"))
    try:
        rows = parse_people_file(upload)
        created, errors = user_store.import_people(rows)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("users_admin"))
    total = created["managers"] + created["teams"]
    if total:
        flash(
            f"Imported {created['managers']} department manager(s) and {created['teams']} team leader(s).",
            "success",
        )
    elif not errors:
        flash("No rows found to import.", "error")
    for message in errors[:12]:
        flash(message, "error")
    if len(errors) > 12:
        flash(f"{len(errors) - 12} more row(s) failed.", "error")
    return redirect(url_for("users_admin"))


@app.route("/users/department/toggle", methods=["POST"])
@hr_required
def users_department_toggle():
    if not csrf_is_valid():
        flash("The form expired. Please refresh and try again.", "error")
        return redirect(url_for("users_admin"))
    try:
        dept_id = int(request.form.get("department_id", "0"))
    except ValueError:
        dept_id = 0
    if user_store.toggle_department(dept_id):
        flash("Department status updated.", "success")
    else:
        flash("Could not update that department.", "error")
    return redirect(url_for("users_admin"))


@app.route("/users/password", methods=["POST"])
@hr_required
def users_password():
    if not csrf_is_valid():
        flash("The form expired. Please refresh and try again.", "error")
        return redirect(url_for("users_admin"))
    try:
        user_id = int(request.form.get("user_id", "0"))
    except ValueError:
        user_id = 0
    password = request.form.get("password", "")
    if len(password) < 8:
        flash("Password must be at least 8 characters.", "error")
    elif user_store.set_password(user_id, password):
        flash("Password updated.", "success")
    else:
        flash("Could not update that password.", "error")
    return redirect(url_for("users_admin"))


@app.route("/users/toggle", methods=["POST"])
@hr_required
def users_toggle():
    if not csrf_is_valid():
        flash("The form expired. Please refresh and try again.", "error")
        return redirect(url_for("users_admin"))
    try:
        user_id = int(request.form.get("user_id", "0"))
    except ValueError:
        user_id = 0
    if user_store.toggle_user(user_id):
        flash("Account status updated.", "success")
    else:
        flash("Could not update that account.", "error")
    return redirect(url_for("users_admin"))


if __name__ == "__main__":
    app.run(debug=False, host="127.0.0.1", port=5000)
